"""
Satisfactory Master Planner - FastAPI backend
"""
import json
import sqlite3
import os
import socket
from contextlib import contextmanager
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from sf_recipes import RECIPES, RAW_RESOURCES
from sf_data import MACHINES, RESOURCES as DEFAULT_RESOURCES, ALT_TIERS, GENERATORS
from solver import RecipeDB, solve, BUDGET_RAWS, TREAT_AS_RAW

# ----- Setup -----
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR.parent / "data"
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = DATA_DIR / "planner.db"

app = FastAPI(title="Satisfactory Master Planner")
HOSTNAME = os.environ.get("PLANNER_HOST", socket.gethostname())
templates = Jinja2Templates(directory=str(BASE_DIR.parent / "frontend"))

db = RecipeDB(RECIPES)
MACHINE_POWER = {m: d["power"] for m, d in MACHINES.items()}


# ----- Database -----
@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db():
    with get_db() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS targets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id INTEGER NOT NULL,
            product TEXT NOT NULL,
            rate_per_min REAL NOT NULL,
            FOREIGN KEY (plan_id) REFERENCES plans(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS recipe_choices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id INTEGER NOT NULL,
            product TEXT NOT NULL,
            recipe TEXT NOT NULL,
            UNIQUE(plan_id, product),
            FOREIGN KEY (plan_id) REFERENCES plans(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS resources (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id INTEGER NOT NULL,
            resource TEXT NOT NULL,
            pure INTEGER DEFAULT 0,
            normal INTEGER DEFAULT 0,
            impure INTEGER DEFAULT 0,
            miner_tier TEXT DEFAULT 'Mk.2',
            UNIQUE(plan_id, resource),
            FOREIGN KEY (plan_id) REFERENCES plans(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS factories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS factory_targets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            factory_id INTEGER NOT NULL,
            product TEXT NOT NULL,
            rate_per_min REAL NOT NULL,
            FOREIGN KEY (factory_id) REFERENCES factories(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS power_plants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id INTEGER NOT NULL,
            name TEXT NOT NULL DEFAULT '',
            generator_type TEXT NOT NULL,
            fuel_type TEXT NOT NULL,
            count INTEGER NOT NULL DEFAULT 1,
            clock_pct REAL NOT NULL DEFAULT 100,
            FOREIGN KEY (plan_id) REFERENCES plans(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS clock_overrides (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id INTEGER NOT NULL,
            product TEXT NOT NULL,
            clock_pct REAL NOT NULL DEFAULT 100,
            UNIQUE(plan_id, product),
            FOREIGN KEY (plan_id) REFERENCES plans(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS factory_choices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            factory_id INTEGER NOT NULL,
            product TEXT NOT NULL,
            recipe TEXT NOT NULL,
            UNIQUE(factory_id, product),
            FOREIGN KEY (factory_id) REFERENCES factories(id) ON DELETE CASCADE
        );
        """)
        # Create a default plan if none exists
        cur = conn.execute("SELECT COUNT(*) c FROM plans")
        if cur.fetchone()["c"] == 0:
            conn.execute("INSERT INTO plans (name) VALUES (?)", ("Default Plan",))
            plan_id = conn.execute("SELECT id FROM plans WHERE name='Default Plan'").fetchone()["id"]
            for res, info in DEFAULT_RESOURCES.items():
                conn.execute(
                    "INSERT INTO resources (plan_id, resource, pure, normal, impure) VALUES (?,?,?,?,?)",
                    (plan_id, res, info["pure"], info["normal"], info["impure"])
                )


init_db()


# ----- Helpers -----
def get_active_plan_id(conn) -> int:
    """For now, always use the first plan. Future: cookie-based selection."""
    row = conn.execute("SELECT id FROM plans ORDER BY id LIMIT 1").fetchone()
    return row["id"] if row else None


# Miner rates at 250% overclock (base * 2.5)
# Mk.1: 60/120/300  Mk.2: 120/240/600  Mk.3: 240/480/1200 (capped at belt speed in-game)
MINER_RATES_PER_PURITY = {
    "Mk.1": {"pure": 300, "normal": 150, "impure": 75},
    "Mk.2": {"pure": 600, "normal": 300, "impure": 150},
    "Mk.3": {"pure": 1200, "normal": 600, "impure": 300},
}
# Oil Extractor / Resource Well at 250% overclock (base 120 * 2.5 = 300 per normal)
EXTRACTOR_RATES = {"pure": 600, "normal": 300, "impure": 150}


def calculate_available(resource: str, pure: int, normal: int, impure: int, tier: str) -> float:
    """Calculate items/min available given node counts and miner tier."""
    if resource == "Water":
        # Water Extractors are effectively unlimited; show a high but finite number
        return 999999
    if resource in ("Crude Oil", "Nitrogen Gas"):
        rates = EXTRACTOR_RATES
    else:
        rates = MINER_RATES_PER_PURITY.get(tier, MINER_RATES_PER_PURITY["Mk.2"])
    return pure * rates["pure"] + normal * rates["normal"] + impure * rates["impure"]


def load_plan_state(plan_id: int) -> dict:
    """Load all of a plan's state from DB."""
    with get_db() as conn:
        plan = conn.execute("SELECT * FROM plans WHERE id=?", (plan_id,)).fetchone()
        targets = conn.execute(
            "SELECT id, product, rate_per_min FROM targets WHERE plan_id=? ORDER BY id",
            (plan_id,)
        ).fetchall()
        choices = conn.execute(
            "SELECT product, recipe FROM recipe_choices WHERE plan_id=?",
            (plan_id,)
        ).fetchall()
        resources = conn.execute(
            "SELECT resource, pure, normal, impure, miner_tier FROM resources WHERE plan_id=? ORDER BY id",
            (plan_id,)
        ).fetchall()
        clocks = conn.execute(
            "SELECT product, clock_pct FROM clock_overrides WHERE plan_id=?",
            (plan_id,)
        ).fetchall()
        pplants = conn.execute(
            "SELECT * FROM power_plants WHERE plan_id=? ORDER BY id",
            (plan_id,)
        ).fetchall()
    return {
        "plan": dict(plan) if plan else None,
        "targets": [dict(t) for t in targets],
        "choices": {c["product"]: c["recipe"] for c in choices},
        "clocks": {c["product"]: c["clock_pct"] for c in clocks},
        "resources": [dict(r) for r in resources],
        "power_plants": [dict(pp) for pp in pplants],
    }


def compute_plan(plan_id: int) -> dict:
    state = load_plan_state(plan_id)
    targets = {t["product"]: t["rate_per_min"] for t in state["targets"]}
    result = solve(targets, state["choices"], db, MACHINE_POWER, state.get("clocks", {}))

    # Also compute a combined result that includes power plant fuel demand
    pp_fuel_targets = {}
    for pp in state.get("power_plants", []):
        gen = GENERATORS.get(pp["generator_type"], {})
        fuel_rate = gen.get("fuels", {}).get(pp["fuel_type"], 0)
        water_rate = gen.get("water_per_min", 0)
        clock = pp["clock_pct"] / 100.0
        total_fuel = fuel_rate * pp["count"] * clock
        total_water = water_rate * pp["count"] * clock
        pp_fuel_targets[pp["fuel_type"]] = pp_fuel_targets.get(pp["fuel_type"], 0) + total_fuel
        if total_water > 0:
            pp_fuel_targets["Water"] = pp_fuel_targets.get("Water", 0) + total_water
    # Merge factory targets + fuel targets for a combined solve
    combined_targets = dict(targets)
    for fuel, rate in pp_fuel_targets.items():
        combined_targets[fuel] = combined_targets.get(fuel, 0) + rate
    combined_result = solve(combined_targets, state["choices"], db, MACHINE_POWER, state.get("clocks", {}))

    # Compute available raw resources
    available = {}
    for r in state["resources"]:
        available[r["resource"]] = calculate_available(
            r["resource"], r["pure"], r["normal"], r["impure"], r["miner_tier"]
        )

    # Build budget table
    budget = []
    for raw in BUDGET_RAWS:
        demand = result["raws"].get(raw, 0)
        avail = available.get(raw, 0)
        surplus = avail - demand
        util = (demand / avail * 100) if avail > 0 else 0
        budget.append({
            "resource": raw,
            "demand": demand,
            "available": avail,
            "surplus": surplus,
            "utilization": util,
            "ok": surplus >= 0,
        })

    # Compute power plant stats
    pp_stats = []
    pp_total_generation = 0
    pp_fuel_demand = {}  # fuel -> total per min needed
    pp_waste_output = {}  # waste product -> total per min produced
    for pp in state.get("power_plants", []):
        gen = GENERATORS.get(pp["generator_type"], {})
        base_mw = gen.get("power_mw", 0)
        fuel_rate = gen.get("fuels", {}).get(pp["fuel_type"], 0)
        water_rate = gen.get("water_per_min", 0)
        waste_rates = gen.get("waste", {}).get(pp["fuel_type"], {})
        clock = pp["clock_pct"] / 100.0
        total_mw = base_mw * pp["count"] * clock
        total_fuel = fuel_rate * pp["count"] * clock
        total_water = water_rate * pp["count"] * clock
        waste_totals = {w: r * pp["count"] * clock for w, r in waste_rates.items()}
        pp_total_generation += total_mw
        pp_fuel_demand[pp["fuel_type"]] = pp_fuel_demand.get(pp["fuel_type"], 0) + total_fuel
        if total_water > 0:
            pp_fuel_demand["Water"] = pp_fuel_demand.get("Water", 0) + total_water
        for w, r in waste_totals.items():
            pp_waste_output[w] = pp_waste_output.get(w, 0) + r
        pp_stats.append({
            **pp,
            "mw_each": base_mw * clock,
            "mw_total": total_mw,
            "fuel_per_min": total_fuel,
            "water_per_min": total_water,
            "waste": waste_totals,
        })

    # Also compute budget for combined chain
    combined_budget = []
    for raw in BUDGET_RAWS:
        demand = combined_result["raws"].get(raw, 0)
        avail = available.get(raw, 0)
        surplus = avail - demand
        util = (demand / avail * 100) if avail > 0 else 0
        combined_budget.append({
            "resource": raw, "demand": demand, "available": avail,
            "surplus": surplus, "utilization": util, "ok": surplus >= 0,
        })

    return {
        "state": state,
        "result": result,
        "combined_result": combined_result,
        "budget": budget,
        "combined_budget": combined_budget,
        "available": available,
        "pp_stats": pp_stats,
        "pp_total_generation": pp_total_generation,
        "pp_fuel_demand": pp_fuel_demand,
        "pp_waste_output": pp_waste_output,
    }


# ----- Routes -----
@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    with get_db() as conn:
        plan_id = get_active_plan_id(conn)
    data = compute_plan(plan_id)

    # Build choices_data: only include products with >1 recipe option
    # (single-recipe products are noise — nothing to choose)
    choices_data = {}
    for product in sorted(db.producers.keys()):
        if product in TREAT_AS_RAW:
            continue
        opts = db.all_choices_for(product)
        if len(opts) < 2:
            continue
        default = db.default_recipe(product)
        choices_data[product] = {
            "default": default,
            "options": [
                {
                    "recipe": rn,
                    "tier": ALT_TIERS.get(rn, ""),
                    "is_alternate": rn.startswith("Alternate:"),
                }
                for rn in opts
            ],
        }

    with get_db() as conn:
        factories_list = load_factories_summary(conn)

    return templates.TemplateResponse("index.html", {
        "request": request,
        "plan_id": plan_id,
        "state": data["state"],
        "result": data["result"],
        "budget": data["budget"],
        "all_products": sorted(db.producers.keys()),
        "choices_data": choices_data,
        "machines_info": MACHINES,
        "factories": factories_list,
        "generators": GENERATORS,
        "pp_stats": data["pp_stats"],
        "pp_total_generation": data["pp_total_generation"],
        "pp_fuel_demand": data["pp_fuel_demand"],
        "pp_waste_output": data["pp_waste_output"],
        "combined_result": data["combined_result"],
        "combined_budget": data["combined_budget"],
        "hostname": HOSTNAME,
    })


@app.post("/api/targets/add")
def add_target(plan_id: int = Form(...), product: str = Form(...), rate: float = Form(...)):
    with get_db() as conn:
        conn.execute(
            "INSERT INTO targets (plan_id, product, rate_per_min) VALUES (?,?,?)",
            (plan_id, product.strip(), rate)
        )
    return RedirectResponse("/", status_code=303)


@app.post("/api/targets/{target_id}/delete")
def delete_target(target_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM targets WHERE id=?", (target_id,))
    return RedirectResponse("/", status_code=303)


@app.post("/api/targets/{target_id}/update")
def update_target(target_id: int, rate: float = Form(...)):
    with get_db() as conn:
        conn.execute("UPDATE targets SET rate_per_min=? WHERE id=?", (rate, target_id))
    return RedirectResponse("/", status_code=303)


@app.post("/api/choices/set")
def set_choice(plan_id: int = Form(...), product: str = Form(...), recipe: str = Form(...)):
    with get_db() as conn:
        # Upsert
        existing = conn.execute(
            "SELECT id FROM recipe_choices WHERE plan_id=? AND product=?",
            (plan_id, product)
        ).fetchone()
        if existing:
            conn.execute("UPDATE recipe_choices SET recipe=? WHERE id=?", (recipe, existing["id"]))
        else:
            conn.execute(
                "INSERT INTO recipe_choices (plan_id, product, recipe) VALUES (?,?,?)",
                (plan_id, product, recipe)
            )
    return RedirectResponse("/#choices", status_code=303)


@app.post("/api/choices/reset")
def reset_choice(plan_id: int = Form(...), product: str = Form(...)):
    with get_db() as conn:
        conn.execute(
            "DELETE FROM recipe_choices WHERE plan_id=? AND product=?",
            (plan_id, product)
        )
    return RedirectResponse("/#choices", status_code=303)


@app.post("/api/resources/update")
def update_resource(resource_id: int = Form(...),
                    pure: int = Form(0),
                    normal: int = Form(0),
                    impure: int = Form(0)):
    with get_db() as conn:
        conn.execute(
            "UPDATE resources SET pure=?, normal=?, impure=? WHERE id=?",
            (pure, normal, impure, resource_id)
        )
    return RedirectResponse("/#resources", status_code=303)


@app.post("/api/resources/tier")
def update_tier(plan_id: int = Form(...), tier: str = Form(...)):
    with get_db() as conn:
        conn.execute("UPDATE resources SET miner_tier=? WHERE plan_id=?", (tier, plan_id))
    return RedirectResponse("/#resources", status_code=303)


@app.post("/api/targets/clear")
def clear_targets(plan_id: int = Form(...)):
    with get_db() as conn:
        conn.execute("DELETE FROM targets WHERE plan_id=?", (plan_id,))
    return RedirectResponse("/", status_code=303)


@app.post("/api/choices/clear")
def clear_choices(plan_id: int = Form(...)):
    with get_db() as conn:
        conn.execute("DELETE FROM recipe_choices WHERE plan_id=?", (plan_id,))
    return RedirectResponse("/#choices", status_code=303)


@app.get("/api/recipe/{product}/options")
def recipe_options(product: str):
    """JSON: list of recipes that produce this product."""
    options = []
    for recipe_name in db.all_choices_for(product):
        rec = db.recipes_by_name[recipe_name]
        out_qty = next((q for o, q in rec["outputs"] if o == product), rec["outputs"][0][1])
        options.append({
            "recipe": recipe_name,
            "machine": rec["machine"],
            "duration": rec["duration"],
            "out_per_min": out_qty * 60 / rec["duration"],
            "tier": ALT_TIERS.get(recipe_name, ""),
            "is_alternate": recipe_name.startswith("Alternate:"),
            "inputs": rec["inputs"],
        })
    return options


@app.get("/api/plan/export")
def export_plan():
    """Export current plan as JSON."""
    with get_db() as conn:
        plan_id = get_active_plan_id(conn)
    state = load_plan_state(plan_id)
    return JSONResponse(state)

@app.get("/api/plan/export/xlsx")
def export_plan_xlsx():
    """Export current plan as a formatted Excel workbook."""
    with get_db() as conn:
        plan_id = get_active_plan_id(conn)
    data = compute_plan(plan_id)
    state = data["state"]
    result = data["result"]
    budget = data["budget"]
    pp_stats = data.get("pp_stats", [])
    pp_total_gen = data.get("pp_total_generation", 0)

    wb = Workbook()

    # ── Styles ──
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B2D42", end_color="2B2D42", fill_type="solid")
    subheader_fill = PatternFill(start_color="3A3D5C", end_color="3A3D5C", fill_type="solid")
    target_fill = PatternFill(start_color="EDF7ED", end_color="EDF7ED", fill_type="solid")
    raw_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    ok_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    warn_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    num_fmt_1 = "0.00"
    num_fmt_pct = r"0.0\%"
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    def style_header(ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def style_data_cell(cell, fmt=None):
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if fmt:
            cell.number_format = fmt

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)

    # ═══════════════════════════════════════════
    # Sheet 1: Production Chain
    # ═══════════════════════════════════════════
    ws = wb.active
    ws.title = "Production Chain"
    headers = ["Product", "Rate /min", "Recipe", "Machine", "# Machines", "Clock %", "Power (MW)", "Target?", "Raw?"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for p in result["products"]:
        row = [
            p["name"],
            p["total_per_min"],
            p.get("recipe") or "—",
            p.get("machine") or "—",
            p.get("machines_ceil") or "",
            p.get("clock_pct"),
            p.get("power_total", 0),
            "Yes" if p.get("is_target") else "",
            "Yes" if p.get("is_raw") else "",
        ]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            style_data_cell(cell)
        # Format numbers
        style_data_cell(ws.cell(row=r, column=2), num_fmt_1)
        style_data_cell(ws.cell(row=r, column=6), num_fmt_1)
        style_data_cell(ws.cell(row=r, column=7), num_fmt_1)
        # Highlight targets and raws
        if p.get("is_target"):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = target_fill
        elif p.get("is_raw"):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = raw_fill

    # Summary row
    ws.append([])
    ws.append(["Total Power (MW)", result.get("total_power", 0)])
    style_data_cell(ws.cell(row=ws.max_row, column=1))
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    style_data_cell(ws.cell(row=ws.max_row, column=2), num_fmt_1)
    auto_width(ws)

    # ═══════════════════════════════════════════
    # Sheet 2: Resource Budget
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("Resource Budget")
    headers2 = ["Resource", "Demand /min", "Available /min", "Surplus /min", "Utilization %", "Status"]
    ws2.append(headers2)
    style_header(ws2, 1, len(headers2))

    for b in budget:
        row = [
            b["resource"],
            b["demand"],
            b["available"],
            b["surplus"],
            b["utilization"],
            "OK" if b["ok"] else "OVER",
        ]
        ws2.append(row)
        r = ws2.max_row
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=r, column=c)
            style_data_cell(cell)
        style_data_cell(ws2.cell(row=r, column=2), num_fmt_1)
        style_data_cell(ws2.cell(row=r, column=3), num_fmt_1)
        style_data_cell(ws2.cell(row=r, column=4), num_fmt_1)
        style_data_cell(ws2.cell(row=r, column=5), num_fmt_pct)
        status_cell = ws2.cell(row=r, column=6)
        status_cell.fill = ok_fill if b["ok"] else warn_fill
        status_cell.font = Font(bold=True, color="1B5E20" if b["ok"] else "B71C1C")
    auto_width(ws2)

    # ═══════════════════════════════════════════
    # Sheet 3: Targets
    # ═══════════════════════════════════════════
    ws3 = wb.create_sheet("Targets")
    headers3 = ["Product", "Rate /min"]
    ws3.append(headers3)
    style_header(ws3, 1, len(headers3))
    for t in state["targets"]:
        ws3.append([t["product"], t["rate_per_min"]])
        r = ws3.max_row
        style_data_cell(ws3.cell(row=r, column=1))
        style_data_cell(ws3.cell(row=r, column=2), num_fmt_1)
    auto_width(ws3)

    # ═══════════════════════════════════════════
    # Sheet 4: Recipe Choices
    # ═══════════════════════════════════════════
    ws4 = wb.create_sheet("Recipe Choices")
    headers4 = ["Product", "Selected Recipe"]
    ws4.append(headers4)
    style_header(ws4, 1, len(headers4))
    for product, recipe in sorted(state["choices"].items()):
        ws4.append([product, recipe])
        r = ws4.max_row
        style_data_cell(ws4.cell(row=r, column=1))
        style_data_cell(ws4.cell(row=r, column=2))
    auto_width(ws4)

    # ═══════════════════════════════════════════
    # Sheet 5: Power Plants (if any)
    # ═══════════════════════════════════════════
    if pp_stats:
        ws5 = wb.create_sheet("Power Plants")
        headers5 = ["Name", "Generator", "Fuel", "Count", "Clock %", "MW Each", "MW Total", "Fuel /min", "Water /min"]
        ws5.append(headers5)
        style_header(ws5, 1, len(headers5))
        for pp in pp_stats:
            row = [
                pp.get("name", ""),
                pp["generator_type"],
                pp["fuel_type"],
                pp["count"],
                pp["clock_pct"],
                pp.get("mw_each", 0),
                pp.get("mw_total", 0),
                pp.get("fuel_per_min", 0),
                pp.get("water_per_min", 0),
            ]
            ws5.append(row)
            r = ws5.max_row
            for c in range(1, len(headers5) + 1):
                style_data_cell(ws5.cell(row=r, column=c))
            style_data_cell(ws5.cell(row=r, column=5), num_fmt_1)
            style_data_cell(ws5.cell(row=r, column=6), num_fmt_1)
            style_data_cell(ws5.cell(row=r, column=7), num_fmt_1)
            style_data_cell(ws5.cell(row=r, column=8), num_fmt_1)
            style_data_cell(ws5.cell(row=r, column=9), num_fmt_1)
        ws5.append([])
        ws5.append(["Total Generation (MW)", "", "", "", "", "", pp_total_gen])
        ws5.cell(row=ws5.max_row, column=1).font = Font(bold=True)
        style_data_cell(ws5.cell(row=ws5.max_row, column=7), num_fmt_1)
        auto_width(ws5)

    # ═══════════════════════════════════════════
    # Sheet 6: Clock Overrides (if any)
    # ═══════════════════════════════════════════
    if state.get("clocks"):
        ws6 = wb.create_sheet("Clock Overrides")
        headers6 = ["Product", "Clock %"]
        ws6.append(headers6)
        style_header(ws6, 1, len(headers6))
        for product, pct in sorted(state["clocks"].items()):
            ws6.append([product, pct])
            r = ws6.max_row
            style_data_cell(ws6.cell(row=r, column=1))
            style_data_cell(ws6.cell(row=r, column=2), num_fmt_1)
        auto_width(ws6)

    # ═══════════════════════════════════════════
    # Sheet 7: Resources (node allocations)
    # ═══════════════════════════════════════════
    if state.get("resources"):
        ws7 = wb.create_sheet("Resource Nodes")
        headers7 = ["Resource", "Pure", "Normal", "Impure", "Miner Tier"]
        ws7.append(headers7)
        style_header(ws7, 1, len(headers7))
        for r_data in state["resources"]:
            ws7.append([r_data["resource"], r_data["pure"], r_data["normal"], r_data["impure"], r_data["miner_tier"]])
            r = ws7.max_row
            for c in range(1, len(headers7) + 1):
                style_data_cell(ws7.cell(row=r, column=c))
        auto_width(ws7)

    # Freeze top row on all sheets
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"

    # Write to buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    plan_name = state["plan"]["name"] if state.get("plan") and state["plan"].get("name") else "plan"
    safe_name = "".join(c for c in plan_name if c.isalnum() or c in " _-").strip().replace(" ", "_")
    filename = f"sf-planner_{safe_name}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )




@app.post("/api/clock/set")
def set_clock(plan_id: int = Form(...), product: str = Form(...), clock_pct: float = Form(...)):
    """Set target clock speed for a product."""
    clock_pct = max(1, min(250, clock_pct))  # clamp 1-250%
    with get_db() as conn:
        existing = conn.execute(
            "SELECT id FROM clock_overrides WHERE plan_id=? AND product=?",
            (plan_id, product)
        ).fetchone()
        if abs(clock_pct - 100.0) < 0.01:
            # 100% is default, just delete the override
            if existing:
                conn.execute("DELETE FROM clock_overrides WHERE id=?", (existing["id"],))
        elif existing:
            conn.execute("UPDATE clock_overrides SET clock_pct=? WHERE id=?", (clock_pct, existing["id"]))
        else:
            conn.execute(
                "INSERT INTO clock_overrides (plan_id, product, clock_pct) VALUES (?,?,?)",
                (plan_id, product, clock_pct)
            )
    return RedirectResponse("/#chain", status_code=303)



# ----- Power plant routes -----
@app.post("/api/powerplants/add")
def add_power_plant(plan_id: int = Form(...),
                    generator_type: str = Form(...),
                    fuel_type: str = Form(...),
                    count: int = Form(1),
                    clock_pct: float = Form(100),
                    pp_name: str = Form("")):
    gen = GENERATORS.get(generator_type)
    if not gen or fuel_type not in gen["fuels"]:
        return RedirectResponse("/", status_code=303)
    clock_pct = max(1, min(250, clock_pct))
    count = max(1, count)
    with get_db() as conn:
        conn.execute(
            "INSERT INTO power_plants (plan_id, name, generator_type, fuel_type, count, clock_pct) VALUES (?,?,?,?,?,?)",
            (plan_id, pp_name.strip(), generator_type, fuel_type, count, clock_pct)
        )
    return RedirectResponse("/#power", status_code=303)


@app.post("/api/powerplants/{pp_id}/update")
def update_power_plant(pp_id: int,
                       count: int = Form(None),
                       clock_pct: float = Form(None),
                       fuel_type: str = Form(None)):
    with get_db() as conn:
        if count is not None:
            conn.execute("UPDATE power_plants SET count=? WHERE id=?", (max(1, count), pp_id))
        if clock_pct is not None:
            conn.execute("UPDATE power_plants SET clock_pct=? WHERE id=?", (max(1, min(250, clock_pct)), pp_id))
        if fuel_type is not None:
            conn.execute("UPDATE power_plants SET fuel_type=? WHERE id=?", (fuel_type, pp_id))
    return RedirectResponse("/#power", status_code=303)


@app.post("/api/powerplants/{pp_id}/delete")
def delete_power_plant(pp_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM power_plants WHERE id=?", (pp_id,))
    return RedirectResponse("/#power", status_code=303)


# ----- Factory helpers -----
def load_factories_summary(conn) -> list:
    """Load all factories with their target summaries."""
    factories = conn.execute(
        "SELECT * FROM factories ORDER BY created_at DESC"
    ).fetchall()
    result = []
    for f in factories:
        targets = conn.execute(
            "SELECT product, rate_per_min FROM factory_targets WHERE factory_id=? ORDER BY id",
            (f["id"],)
        ).fetchall()
        choices_count = conn.execute(
            "SELECT COUNT(*) c FROM factory_choices WHERE factory_id=?",
            (f["id"],)
        ).fetchone()["c"]
        # Load power plants for this factory's plan (via plan_id reference)
        result.append({
            "id": f["id"],
            "name": f["name"],
            "created_at": f["created_at"],
            "targets": [dict(t) for t in targets],
            "choices_count": choices_count,
        })
    return result


# ----- Factory routes -----
@app.post("/api/factories/save")
def save_factory(plan_id: int = Form(...), factory_name: str = Form(...)):
    """Snapshot the current plan targets + choices into a named factory."""
    if not factory_name.strip():
        return RedirectResponse("/", status_code=303)
    with get_db() as conn:
        conn.execute("INSERT INTO factories (name) VALUES (?)", (factory_name.strip(),))
        fid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        targets = conn.execute(
            "SELECT product, rate_per_min FROM targets WHERE plan_id=?", (plan_id,)
        ).fetchall()
        for t in targets:
            conn.execute(
                "INSERT INTO factory_targets (factory_id, product, rate_per_min) VALUES (?,?,?)",
                (fid, t["product"], t["rate_per_min"])
            )
        choices = conn.execute(
            "SELECT product, recipe FROM recipe_choices WHERE plan_id=?", (plan_id,)
        ).fetchall()
        for c in choices:
            conn.execute(
                "INSERT INTO factory_choices (factory_id, product, recipe) VALUES (?,?,?)",
                (fid, c["product"], c["recipe"])
            )
    return RedirectResponse("/", status_code=303)


@app.post("/api/factories/{factory_id}/delete")
def delete_factory(factory_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM factories WHERE id=?", (factory_id,))
    return RedirectResponse("/", status_code=303)


@app.post("/api/factories/{factory_id}/load")
def load_factory(factory_id: int):
    """Load a factory back into the active plan for editing."""
    with get_db() as conn:
        plan_id = get_active_plan_id(conn)
        factory = conn.execute("SELECT * FROM factories WHERE id=?", (factory_id,)).fetchone()
        if not factory:
            return RedirectResponse("/", status_code=303)
        conn.execute("DELETE FROM targets WHERE plan_id=?", (plan_id,))
        conn.execute("DELETE FROM recipe_choices WHERE plan_id=?", (plan_id,))
        ftargets = conn.execute(
            "SELECT product, rate_per_min FROM factory_targets WHERE factory_id=?",
            (factory_id,)
        ).fetchall()
        for t in ftargets:
            conn.execute(
                "INSERT INTO targets (plan_id, product, rate_per_min) VALUES (?,?,?)",
                (plan_id, t["product"], t["rate_per_min"])
            )
        fchoices = conn.execute(
            "SELECT product, recipe FROM factory_choices WHERE factory_id=?",
            (factory_id,)
        ).fetchall()
        for c in fchoices:
            conn.execute(
                "INSERT INTO recipe_choices (plan_id, product, recipe) VALUES (?,?,?)",
                (plan_id, c["product"], c["recipe"])
            )
    return RedirectResponse("/", status_code=303)


@app.get("/world", response_class=HTMLResponse)
def world_view(request: Request):
    """Aggregate all factories + power plants by solving each individually and summing results."""
    with get_db() as conn:
        plan_id = get_active_plan_id(conn)
        factories_list = load_factories_summary(conn)

        # Load plan-level resources, choices, clocks
        plan_choices = {c["product"]: c["recipe"] for c in conn.execute(
            "SELECT product, recipe FROM recipe_choices WHERE plan_id=?", (plan_id,)
        ).fetchall()}
        plan_clocks = {c["product"]: c["clock_pct"] for c in conn.execute(
            "SELECT product, clock_pct FROM clock_overrides WHERE plan_id=?", (plan_id,)
        ).fetchall()}
        resources = conn.execute(
            "SELECT resource, pure, normal, impure, miner_tier FROM resources WHERE plan_id=? ORDER BY id",
            (plan_id,)
        ).fetchall()

        # Solve each factory individually
        world_raws = {}  # raw -> total demand
        world_power = 0
        world_products = {}  # product -> {total_per_min, machines, power, ...}
        factory_details = []

        for f in factories_list:
            ftargets = {t["product"]: t["rate_per_min"] for t in f["targets"]}
            fchoices = {c["product"]: c["recipe"] for c in conn.execute(
                "SELECT product, recipe FROM factory_choices WHERE factory_id=?",
                (f["id"],)
            ).fetchall()}
            fresult = solve(ftargets, fchoices, db, MACHINE_POWER, plan_clocks)
            factory_details.append({"name": f["name"], "power": fresult["total_power"], "targets": f["targets"]})
            world_power += fresult["total_power"]
            for raw, rate in fresult["raws"].items():
                world_raws[raw] = world_raws.get(raw, 0) + rate
            for p in fresult["products"]:
                if p["name"] in world_products:
                    wp = world_products[p["name"]]
                    wp["total_per_min"] += p["total_per_min"]
                    wp["power_total"] += p["power_total"]
                    if p["machines_ceil"]:
                        wp["machines_ceil"] = (wp["machines_ceil"] or 0) + p["machines_ceil"]
                else:
                    world_products[p["name"]] = {**p}

        # Solve each power plant individually
        pplants = conn.execute(
            "SELECT * FROM power_plants WHERE plan_id=? ORDER BY id", (plan_id,)
        ).fetchall()
        pp_total_gen = 0
        world_waste = {}  # produced by generators
        world_waste_consumed = {}  # consumed by fuel chains
        pp_details = []

        for pp in pplants:
            pp = dict(pp)
            gen = GENERATORS.get(pp["generator_type"], {})
            base_mw = gen.get("power_mw", 0)
            fuel_rate = gen.get("fuels", {}).get(pp["fuel_type"], 0)
            water_rate = gen.get("water_per_min", 0)
            waste_rates = gen.get("waste", {}).get(pp["fuel_type"], {})
            clock = pp["clock_pct"] / 100.0
            total_mw = base_mw * pp["count"] * clock
            total_fuel = fuel_rate * pp["count"] * clock
            total_water = water_rate * pp["count"] * clock
            pp_total_gen += total_mw

            for w, r in waste_rates.items():
                world_waste[w] = world_waste.get(w, 0) + r * pp["count"] * clock

            # Solve the fuel chain for this power plant
            pp_targets = {pp["fuel_type"]: total_fuel}
            if total_water > 0:
                pp_targets["Water"] = total_water
            pp_result = solve(pp_targets, plan_choices, db, MACHINE_POWER, plan_clocks)

            pp_details.append({
                "id": pp["id"], "name": pp.get("name", ""),
                "generator_type": pp["generator_type"],
                "fuel_type": pp["fuel_type"],
                "count": pp["count"], "clock_pct": pp["clock_pct"],
                "mw_total": total_mw,
                "fuel_per_min": total_fuel,
                "chain_power": pp_result["total_power"],
            })

            world_power += pp_result["total_power"]
            for raw, rate in pp_result["raws"].items():
                world_raws[raw] = world_raws.get(raw, 0) + rate
                # Track waste consumed by fuel chains separately
                if raw in ("Uranium Waste", "Plutonium Waste"):
                    world_waste_consumed[raw] = world_waste_consumed.get(raw, 0) + rate
            for p in pp_result["products"]:
                if p["name"] in world_products:
                    wp = world_products[p["name"]]
                    wp["total_per_min"] += p["total_per_min"]
                    wp["power_total"] += p["power_total"]
                    if p["machines_ceil"]:
                        wp["machines_ceil"] = (wp["machines_ceil"] or 0) + p["machines_ceil"]
                else:
                    world_products[p["name"]] = {**p}

    # Build sorted products list (intermediates first, then raws)
    products_list = sorted(world_products.values(),
                          key=lambda p: (1 if p["is_raw"] else 0, p["name"]))

    # Offset waste demand with waste supply from generators.
    # If factories consume Uranium Waste (e.g. to make Pu Fuel Rods),
    # and power plants produce it, the net demand is reduced.
    waste_balance = {}
    for waste_type, produced in world_waste.items():
        consumed_by_chains = world_waste_consumed.get(waste_type, 0)
        consumed_by_factories = world_raws.get(waste_type, 0)
        total_consumed = consumed_by_chains + consumed_by_factories
        net = produced - total_consumed
        waste_balance[waste_type] = {
            "produced": produced,
            "consumed_by_pp_chains": consumed_by_chains,
            "consumed_by_factories": consumed_by_factories,
            "total_consumed": total_consumed,
            "net": net,
        }
        # Remove waste from world_raws since it's supplied by generators, not mined
        if waste_type in world_raws:
            # If supply >= demand, no shortfall. If not, show deficit.
            supplied = min(produced, consumed_by_factories + consumed_by_chains)
            remaining_demand = max(0, (consumed_by_factories + consumed_by_chains) - produced)
            # Don't show waste in the mining budget at all - it's not a mineable
            if waste_type in world_raws:
                del world_raws[waste_type]

    # Build budget from summed raws
    available = {}
    for r in resources:
        available[r["resource"]] = calculate_available(
            r["resource"], r["pure"], r["normal"], r["impure"], r["miner_tier"]
        )

    budget = []
    for raw in BUDGET_RAWS:
        demand = world_raws.get(raw, 0)
        avail = available.get(raw, 0)
        surplus = avail - demand
        util = (demand / avail * 100) if avail > 0 else 0
        budget.append({
            "resource": raw, "demand": demand, "available": avail,
            "surplus": surplus, "utilization": util, "ok": surplus >= 0,
        })

    # Calculate sink: factory production of fuel items vs PP consumption
    # Factory targets that match PP fuel types represent intentional overproduction
    pp_fuel_consumption = {}  # fuel_type -> total consumed/min by power plants
    for ppd in pp_details:
        fuel = ppd["fuel_type"]
        pp_fuel_consumption[fuel] = pp_fuel_consumption.get(fuel, 0) + ppd["fuel_per_min"]

    factory_fuel_production = {}  # product -> total targeted/min by factories
    for f in factories_list:
        for t in f["targets"]:
            if t["product"] in pp_fuel_consumption:
                factory_fuel_production[t["product"]] = factory_fuel_production.get(t["product"], 0) + t["rate_per_min"]

    sink_items = {}
    for product, produced in factory_fuel_production.items():
        consumed = pp_fuel_consumption.get(product, 0)
        surplus = produced - consumed
        if surplus > 0.01:
            sink_items[product] = {
                "produced": produced,
                "consumed": consumed,
                "sunk": surplus,
            }

    return templates.TemplateResponse("world.html", {
        "request": request,
        "factories": factories_list,
        "factory_details": factory_details,
        "pp_details": pp_details,
        "sink_items": sink_items,
        "result": {"products": products_list, "total_power": world_power},
        "budget": budget,
        "total_power": world_power,
        "pp_total_generation": pp_total_gen,
        "world_waste": world_waste,
        "world_waste_consumed": world_waste_consumed,
        "waste_balance": waste_balance,
        "machines_info": MACHINES,
        "hostname": HOSTNAME,
    })






@app.get("/factory/{factory_id}", response_class=HTMLResponse)
def factory_detail(request: Request, factory_id: int):
    """Show a saved factory with its full production chain."""
    with get_db() as conn:
        plan_id = get_active_plan_id(conn)
        factory = conn.execute("SELECT * FROM factories WHERE id=?", (factory_id,)).fetchone()
        if not factory:
            return RedirectResponse("/", status_code=303)
        factory = dict(factory)

        # Load factory targets and choices
        ftargets = conn.execute(
            "SELECT product, rate_per_min FROM factory_targets WHERE factory_id=? ORDER BY id",
            (factory_id,)
        ).fetchall()
        fchoices = conn.execute(
            "SELECT product, recipe FROM factory_choices WHERE factory_id=?",
            (factory_id,)
        ).fetchall()
        choices_dict = {c["product"]: c["recipe"] for c in fchoices}

        # Load plan-level clocks and resources
        plan_clocks = {c["product"]: c["clock_pct"] for c in conn.execute(
            "SELECT product, clock_pct FROM clock_overrides WHERE plan_id=?", (plan_id,)
        ).fetchall()}
        resources = conn.execute(
            "SELECT resource, pure, normal, impure, miner_tier FROM resources WHERE plan_id=? ORDER BY id",
            (plan_id,)
        ).fetchall()

    # Solve
    targets = {t["product"]: t["rate_per_min"] for t in ftargets}
    result = solve(targets, choices_dict, db, MACHINE_POWER, plan_clocks)

    # Budget
    available = {}
    for r in resources:
        available[r["resource"]] = calculate_available(
            r["resource"], r["pure"], r["normal"], r["impure"], r["miner_tier"]
        )
    budget = []
    for raw in BUDGET_RAWS:
        demand = result["raws"].get(raw, 0)
        avail = available.get(raw, 0)
        surplus = avail - demand
        util = (demand / avail * 100) if avail > 0 else 0
        budget.append({
            "resource": raw, "demand": demand, "available": avail,
            "surplus": surplus, "utilization": util, "ok": surplus >= 0,
        })

    # Choices data for recipe picker
    choices_data = {}
    for product in sorted(db.producers.keys()):
        if product in TREAT_AS_RAW:
            continue
        opts = db.all_choices_for(product)
        if len(opts) < 2:
            continue
        default = db.default_recipe(product)
        choices_data[product] = {
            "default": default,
            "options": [
                {"recipe": rn, "tier": ALT_TIERS.get(rn, ""), "is_alternate": rn.startswith("Alternate:")}
                for rn in opts
            ],
        }

    return templates.TemplateResponse("factory_detail.html", {
        "request": request,
        "plan_id": plan_id,
        "factory": factory,
        "targets": [dict(t) for t in ftargets],
        "result": result,
        "budget": budget,
        "choices_data": choices_data,
        "choices_dict": choices_dict,
        "machines_info": MACHINES,
        "hostname": HOSTNAME,
    })


@app.post("/api/clock/set/pp/{pp_id}")
def set_clock_for_pp(pp_id: int, plan_id: int = Form(...), product: str = Form(...), clock_pct: float = Form(...)):
    """Set clock override and redirect back to power plant detail."""
    clock_pct = max(1, min(250, clock_pct))
    with get_db() as conn:
        existing = conn.execute(
            "SELECT id FROM clock_overrides WHERE plan_id=? AND product=?",
            (plan_id, product)
        ).fetchone()
        if abs(clock_pct - 100.0) < 0.01:
            if existing:
                conn.execute("DELETE FROM clock_overrides WHERE id=?", (existing["id"],))
        elif existing:
            conn.execute("UPDATE clock_overrides SET clock_pct=? WHERE id=?", (clock_pct, existing["id"]))
        else:
            conn.execute(
                "INSERT INTO clock_overrides (plan_id, product, clock_pct) VALUES (?,?,?)",
                (plan_id, product, clock_pct)
            )
    return RedirectResponse(f"/powerplant/{pp_id}#chain", status_code=303)


@app.get("/powerplant/{pp_id}", response_class=HTMLResponse)
def powerplant_detail(request: Request, pp_id: int):
    """Show a single power plant with its full fuel production chain."""
    with get_db() as conn:
        plan_id = get_active_plan_id(conn)
        pp = conn.execute("SELECT * FROM power_plants WHERE id=?", (pp_id,)).fetchone()
        if not pp:
            return RedirectResponse("/", status_code=303)
        pp = dict(pp)

        # Load plan choices and clocks for recipe/clock overrides
        choices = conn.execute(
            "SELECT product, recipe FROM recipe_choices WHERE plan_id=?", (plan_id,)
        ).fetchall()
        choices_dict = {c["product"]: c["recipe"] for c in choices}
        clocks = conn.execute(
            "SELECT product, clock_pct FROM clock_overrides WHERE plan_id=?", (plan_id,)
        ).fetchall()
        clocks_dict = {c["product"]: c["clock_pct"] for c in clocks}

        # Load resources for budget
        resources = conn.execute(
            "SELECT resource, pure, normal, impure, miner_tier FROM resources WHERE plan_id=? ORDER BY id",
            (plan_id,)
        ).fetchall()

    # Compute fuel demand for this power plant
    gen = GENERATORS.get(pp["generator_type"], {})
    base_mw = gen.get("power_mw", 0)
    fuel_rate = gen.get("fuels", {}).get(pp["fuel_type"], 0)
    water_rate = gen.get("water_per_min", 0)
    clock = pp["clock_pct"] / 100.0
    total_mw = base_mw * pp["count"] * clock
    total_fuel = fuel_rate * pp["count"] * clock
    total_water = water_rate * pp["count"] * clock

    # Compute waste output
    waste_rates = gen.get("waste", {}).get(pp["fuel_type"], {})
    waste_totals = {w: r * pp["count"] * clock for w, r in waste_rates.items()}

    # Build targets from fuel demand
    targets = {pp["fuel_type"]: total_fuel}
    if total_water > 0:
        targets["Water"] = total_water

    # Solve the production chain
    result = solve(targets, choices_dict, db, MACHINE_POWER, clocks_dict)

    # Compute resource budget
    available = {}
    for r in resources:
        available[r["resource"]] = calculate_available(
            r["resource"], r["pure"], r["normal"], r["impure"], r["miner_tier"]
        )
    budget = []
    for raw in BUDGET_RAWS:
        demand = result["raws"].get(raw, 0)
        avail = available.get(raw, 0)
        surplus = avail - demand
        util = (demand / avail * 100) if avail > 0 else 0
        budget.append({
            "resource": raw, "demand": demand, "available": avail,
            "surplus": surplus, "utilization": util, "ok": surplus >= 0,
        })

    # Build choices_data for recipe overrides display
    choices_data = {}
    for product in sorted(db.producers.keys()):
        if product in TREAT_AS_RAW:
            continue
        opts = db.all_choices_for(product)
        if len(opts) < 2:
            continue
        default = db.default_recipe(product)
        choices_data[product] = {
            "default": default,
            "options": [
                {"recipe": rn, "tier": ALT_TIERS.get(rn, ""), "is_alternate": rn.startswith("Alternate:")}
                for rn in opts
            ],
        }

    pp_info = {
        **pp,
        "mw_total": total_mw,
        "fuel_per_min": total_fuel,
        "water_per_min": total_water,
        "waste": waste_totals,
    }

    return templates.TemplateResponse("powerplant_detail.html", {
        "request": request,
        "plan_id": plan_id,
        "pp": pp_info,
        "result": result,
        "budget": budget,
        "choices_data": choices_data,
        "choices_dict": choices_dict,
        "machines_info": MACHINES,
        "generators": GENERATORS,
        "hostname": HOSTNAME,
    })


@app.get("/power", response_class=HTMLResponse)
def power_budget(request: Request):
    """Power budget: all factories consumption vs all power plants generation."""
    with get_db() as conn:
        plan_id = get_active_plan_id(conn)
        factories_list = load_factories_summary(conn)

        # Get power plants from active plan
        pplants = conn.execute(
            "SELECT * FROM power_plants WHERE plan_id=? ORDER BY id", (plan_id,)
        ).fetchall()

    # Compute factory power consumption per factory
    factory_power = []
    total_consumption = 0
    for f in factories_list:
        # Solve each factory individually to get its power consumption
        ftargets = {t["product"]: t["rate_per_min"] for t in f["targets"]}
        # Load factory choices
        with get_db() as conn:
            fchoices_rows = conn.execute(
                "SELECT product, recipe FROM factory_choices WHERE factory_id=?",
                (f["id"],)
            ).fetchall()
        fchoices = {c["product"]: c["recipe"] for c in fchoices_rows}
        fresult = solve(ftargets, fchoices, db, MACHINE_POWER)
        factory_power.append({
            "name": f["name"],
            "power_mw": fresult["total_power"],
            "targets": f["targets"],
        })
        total_consumption += fresult["total_power"]

    # Compute power plant generation
    pp_stats = []
    total_generation = 0
    pp_fuel_demand = {}
    for pp in pplants:
        pp = dict(pp)
        gen = GENERATORS.get(pp["generator_type"], {})
        base_mw = gen.get("power_mw", 0)
        fuel_rate = gen.get("fuels", {}).get(pp["fuel_type"], 0)
        water_rate = gen.get("water_per_min", 0)
        clock = pp["clock_pct"] / 100.0
        total_mw = base_mw * pp["count"] * clock
        total_fuel = fuel_rate * pp["count"] * clock
        total_water = water_rate * pp["count"] * clock
        total_generation += total_mw
        pp_fuel_demand[pp["fuel_type"]] = pp_fuel_demand.get(pp["fuel_type"], 0) + total_fuel
        if total_water > 0:
            pp_fuel_demand["Water"] = pp_fuel_demand.get("Water", 0) + total_water
        pp_stats.append({
            **pp,
            "mw_each": base_mw * clock,
            "mw_total": total_mw,
            "fuel_per_min": total_fuel,
        })

    return templates.TemplateResponse("power.html", {
        "request": request,
        "factory_power": factory_power,
        "total_consumption": total_consumption,
        "pp_stats": pp_stats,
        "total_generation": total_generation,
        "net_power": total_generation - total_consumption,
        "pp_fuel_demand": pp_fuel_demand,
        "generators": GENERATORS,
        "hostname": HOSTNAME,
    })


@app.get("/health")
def health():
    return {"status": "ok"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)


# ----- Catch-all: redirect stray GETs on POST-only API routes back home -----
@app.get("/api/{path:path}")
def api_catch_all(path: str):
    return RedirectResponse("/", status_code=303)
